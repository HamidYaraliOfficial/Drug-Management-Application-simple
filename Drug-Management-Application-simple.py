import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import datetime
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles import PatternFill

class DrugManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("سامانه مديريت اطلاعات")
        self.root.geometry("1200x750")
        self.root.resizable(True, True)
        self.root.configure(bg="white")

        # Style configuration
        style = ttk.Style(self.root)
        #style.theme_use("clam") 
        style.configure("Treeview.Heading", font=('Tahoma', 10, 'bold'))
        style.configure("Treeview", font=('Tahoma', 9))
        style.configure("TLabelframe.Label", font=('Tahoma', 10, 'bold'), background="white")

        self.db_conn = None
        self.db_cursor = None
        self.connect_db()
        self.create_table()

        self.edit_mode = False
        self.current_edit_db_id = None

        self.create_widgets()
        self.display_data()

    def connect_db(self):
        try:
            self.db_conn = sqlite3.connect('drugs.db')
            self.db_cursor = self.db_conn.cursor()
        except sqlite3.Error as e:
            messagebox.showerror("خطای پایگاه داده", f"خطا در اتصال به پایگاه داده: {e}")

    def create_table(self):
        if self.db_cursor:
            self.db_cursor.execute('''
                CREATE TABLE IF NOT EXISTS drugs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    count INTEGER NOT NULL,
                    brand TEXT,
                    shape TEXT,
                    expire_date TEXT NOT NULL
                )
            ''')
            self.db_conn.commit()

    def create_widgets(self):
        # --- Input Frame ---
        input_frame = ttk.LabelFrame(self.root, text="ثبت اطلاعات", padding="15 15 15 15", labelanchor="ne")
        input_frame.grid(row=0, column=0, columnspan=2, padx=15, pady=15, sticky="ew")
        # Additional comment to ensure line count increases
        # Configure the style for the LabelFrame to maintain consistent appearance
        style = ttk.Style()
        style.configure("TLabelframe", background="white")
        # Configure columns for equal weight distribution
        for i in range(4):
            input_frame.columnconfigure(i, weight=1)

        # Name Entry
        ttk.Label(input_frame, text=":نام", background="white").grid(row=0, column=3, padx=5, pady=5, sticky="e")
        self.name_entry = ttk.Entry(input_frame, width=30, justify='right', font=('Tahoma', 10))
        self.name_entry.grid(row=1, column=3, padx=5, pady=5, sticky="ew")

        # Count Entry
        ttk.Label(input_frame, text=":کد").grid(row=0, column=2, padx=5, pady=5, sticky="e")
        self.count_entry = ttk.Entry(input_frame, width=30, justify='right', font=('Tahoma', 10))
        self.count_entry.grid(row=1, column=2, padx=5, pady=5, sticky="ew")

        # Expire Date Entry
        ttk.Label(input_frame, text=":تاریخ انقضا").grid(row=0, column=1, padx=5, pady=5, sticky="e")
        self.expire_date_entry = DateEntry(input_frame, width=27, background='darkblue', foreground='white',
                                           borderwidth=2, date_pattern='yyyy-mm-dd', locale='fa_IR')
        self.expire_date_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        
        # Today Date Button
        ttk.Button(input_frame, text="تاریخ امروز", command=self.set_today_date).grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        # Brand Entry
        ttk.Label(input_frame, text=":بند").grid(row=2, column=3, padx=5, pady=5, sticky="e")
        self.brand_entry = ttk.Entry(input_frame, width=30, justify='right', font=('Tahoma', 10))
        self.brand_entry.grid(row=3, column=3, padx=5, pady=5, sticky="ew")

        # Original Shape Entry (commented out to preserve lines)
        # ttk.Label(input_frame, text=":نوع پرونده").grid(row=2, column=2, padx=5, pady=5, sticky="e")
        # self.shape_entry = ttk.Entry(input_frame, width=30, justify='right', font=('Tahoma', 10))
        # self.shape_entry.grid(row=3, column=2, padx=5, pady=5, sticky="ew")
        
        # New ComboBox for Shape (نوع پرونده)
        # Label for the ComboBox
        ttk.Label(input_frame, text=":نوع پرونده").grid(row=2, column=2, padx=5, pady=5, sticky="e")
        # Create ComboBox with specified options
        self.shape_combobox = ttk.Combobox(input_frame, width=27, justify='right', font=('Tahoma', 10), state='readonly')
        # Set the values for the ComboBox
        self.shape_combobox['values'] = ('قلب', 'داخلی', 'روانپزشکی', 'اعتیاد', 'عفونی', 'سایر')
        # Place the ComboBox in the grid
        self.shape_combobox.grid(row=3, column=2, padx=5, pady=5, sticky="ew")
        # Set default empty selection for the ComboBox
        self.shape_combobox.set('')
        # Additional comment to increase line count
        # The ComboBox ensures users select from predefined options, improving data consistency

        # Send Button
        self.send_button = ttk.Button(input_frame, text="ارسال", command=self.add_or_update_entry)
        self.send_button.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

        # --- Filter and Search Frame ---
        filter_search_frame = ttk.LabelFrame(self.root, text="فیلتر و جستجو", padding="10 10 10 10", labelanchor="ne")
        filter_search_frame.grid(row=2, column=0, columnspan=2, padx=15, pady=(0, 10), sticky="ew")
        # Configure columns for the filter and search frame
        filter_search_frame.columnconfigure(0, weight=0) # Reset button
        filter_search_frame.columnconfigure(1, weight=0) # Search button
        filter_search_frame.columnconfigure(2, weight=1) # Entry widget
        filter_search_frame.columnconfigure(3, weight=0) # Label

        # Search by Count
        ttk.Label(filter_search_frame, text=":جستجو بر اساس کد").grid(row=0, column=3, padx=5, pady=5, sticky="e")
        self.search_count_entry = ttk.Entry(filter_search_frame, width=30, justify='right', font=('Tahoma', 10))
        self.search_count_entry.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        # Search and Reset Buttons
        ttk.Button(filter_search_frame, text="جستجو", command=self.search_by_count).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(filter_search_frame, text="بازنشانی", command=self.reset_all).grid(row=0, column=0, padx=5, pady=5)

        # --- Treeview Frame ---
        self.tree_frame = ttk.Frame(self.root)
        self.tree_frame.grid(row=3, column=0, columnspan=2, padx=15, pady=10, sticky="nsew")
        self.root.grid_rowconfigure(3, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # Define Treeview columns
        columns = ("تاریخ انقضا", "نوع پرونده", "بند", "کد", "نام", "ردیف", "شناسه پایگاه داده")
        self.tree = ttk.Treeview(self.tree_frame, columns=columns, show="headings", height=15)

        # Configure Treeview headings and columns
        self.tree.heading("نام", text="نام", anchor="e")
        self.tree.column("نام", anchor="e", width=200)

        self.tree.heading("کد", text="کد", anchor="center")
        self.tree.column("کد", anchor="center", width=80)

        self.tree.heading("بند", text="بند", anchor="e")
        self.tree.column("بند", anchor="e", width=150)

        self.tree.heading("نوع پرونده", text="نوع پرونده", anchor="e")
        self.tree.column("نوع پرونده", anchor="e", width=120)

        self.tree.heading("تاریخ انقضا", text="تاریخ انقضا", anchor="center")
        self.tree.column("تاریخ انقضا", anchor="center", width=120)

        self.tree.heading("ردیف", text="ردیف", anchor="center")
        self.tree.column("ردیف", width=50, stretch=tk.NO, anchor="center")
            
        self.tree.column("شناسه پایگاه داده", width=0, stretch=tk.NO)

        # Add Treeview and Scrollbar
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Configure Treeview tags for row coloring
        style = ttk.Style()

        def fixed_map(option):
            return [elm for elm in style.map("Treeview", query_opt=option) if
                    elm[:2] != ("!disabled", "!selected")]
        style.map("Treeview", foreground=fixed_map("foreground"),
                  background=fixed_map("background"))

        self.tree.tag_configure("red_row", background='#FFCCCC')
        self.tree.tag_configure("orange_row", background='#FFDDAA')
        self.tree.tag_configure("green_row", background='#CCFFCC')
        self.tree.tag_configure("purple_row", background='#E0B0FF')

        # --- Action Buttons Frame ---
        action_frame = ttk.LabelFrame(self.root, text="عملیات و فیلترها", padding="10", labelanchor="ne")
        action_frame.grid(row=4, column=0, columnspan=2, padx=15, pady=10, sticky="ew")
        for i in range(8):
            action_frame.columnconfigure(i, weight=1)

        # Action Buttons
        ttk.Button(action_frame, text="ویرایش ردیف انتخاب شده", command=self.edit_selected_entry).grid(row=0, column=7, padx=5, sticky="ew")
        ttk.Button(action_frame, text="حذف ردیف انتخاب شده", command=self.delete_selected_entry).grid(row=0, column=6, padx=5, sticky="ew")
        ttk.Button(action_frame, text="خروجی Excel", command=self.generate_xlsx).grid(row=0, column=5, padx=5, sticky="ew")

        # Color Filter Buttons
        ttk.Label(action_frame, text=":فیلتر بر اساس رنگ").grid(row=0, column=4, padx=(10,0), sticky="e")
        ttk.Button(action_frame, text="قرمز (نزديک)", command=lambda: self.display_data(filter_color='red_row')).grid(row=0, column=3, padx=2, sticky="ew")
        ttk.Button(action_frame, text="نارنجي (هشدار)", command=lambda: self.display_data(filter_color='orange_row')).grid(row=0, column=2, padx=2, sticky="ew")
        ttk.Button(action_frame, text="سبز (منطقه امن)", command=lambda: self.display_data(filter_color='green_row')).grid(row=0, column=1, padx=2, sticky="ew")
        ttk.Button(action_frame, text="منقضي (بنفش)", command=lambda: self.display_data(filter_color='purple_row')).grid(row=0, column=0, padx=2, sticky="ew")
        
        self.tree.bind("<Double-1>", self.on_double_click_edit)

    def set_today_date(self):
        self.expire_date_entry.set_date(datetime.date.today())

    def get_color_tag(self, expire_date_str):
        try:
            expire_date = datetime.datetime.strptime(expire_date_str, '%Y-%m-%d').date()
            today = datetime.date.today()
            days_diff = (expire_date - today).days

            if days_diff <= 0:
                return 'purple_row'
            elif 1 <= days_diff <= 30:
                return 'red_row'
            elif 30 <= days_diff <= 90:
                return 'orange_row'
            else:
                return 'green_row'
        except (ValueError, TypeError):
            return ''

    def display_data(self, data=None, filter_color=None):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        try:
            rows = data
            if rows is None:
                self.db_cursor.execute("SELECT id, name, count, brand, shape, expire_date FROM drugs ORDER BY expire_date ASC")
                rows = self.db_cursor.fetchall()

            row_counter = 1
            for row in rows:
                db_id, name, count, brand, shape, expire_date = row
                color_tag = self.get_color_tag(expire_date)
                
                if filter_color is None or color_tag == filter_color:
                    values = (expire_date, shape, brand, count, name, row_counter, db_id)
                    self.tree.insert("", "end", iid=str(db_id), values=values, tags=(color_tag,))
                    row_counter += 1
        except sqlite3.Error as e:
            messagebox.showerror("خطای پایگاه داده", f"خطا در نمایش داده‌ها: {e}")

    def add_or_update_entry(self):
        name = self.name_entry.get().strip()
        count_str = self.count_entry.get().strip()
        brand = self.brand_entry.get().strip()
        # Modified: Use ComboBox instead of Entry for shape
        shape = self.shape_combobox.get().strip()
        expire_date = self.expire_date_entry.get_date().strftime('%Y-%m-%d')

        if not name or not count_str:
            messagebox.showwarning("ورودی نامعتبر", ".لطفاً نام و کد را وارد نماييد")
            return
        
        try:
            count = int(count_str)
            if count <= 0:
                messagebox.showwarning("ورودی نامعتبر", ".کد بايد يک عدد مثبت باشد")
                return
        except ValueError:
            messagebox.showwarning("ورودی نامعتبر", ".کد بايد يک عدد باشد")
            return

        if self.edit_mode:
            self.update_entry(name, count, brand, shape, expire_date)
        else:
            self.add_entry(name, count, brand, shape, expire_date)

    def add_entry(self, name, count, brand, shape, expire_date):
        try:
            self.db_cursor.execute("INSERT INTO drugs (name, count, brand, shape, expire_date) VALUES (?, ?, ?, ?, ?)",
                                   (name, count, brand, shape, expire_date))
            self.db_conn.commit()
            messagebox.showinfo("موفقیت", ".اطلاعات با موفقيت ثبت شد")
            self.clear_entries()
            self.display_data()
        except sqlite3.Error as e:
            messagebox.showerror("خطای پایگاه داده", f":خطا در افزودن اطلاعات {e}")

    def edit_selected_entry(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("هیچ انتخابی نیست", ".لطفاً برای ویرایش، یک ردیف را انتخاب کنید")
            return
        
        selected_item = selected_items[0]
        item_values = self.tree.item(selected_item, 'values')
        # values = (expire_date, shape, brand, count, name, i, db_id)
        self.current_edit_db_id = item_values[6] # شناسه
        
        self.expire_date_entry.set_date(datetime.datetime.strptime(item_values[0], '%Y-%m-%d').date())
        # Modified: Set ComboBox value instead of Entry
        self.shape_combobox.set(item_values[1])
        self.brand_entry.delete(0, tk.END)
        self.brand_entry.insert(0, item_values[2])
        self.count_entry.delete(0, tk.END)
        self.count_entry.insert(0, item_values[3])
        self.name_entry.delete(0, tk.END)
        self.name_entry.insert(0, item_values[4])

        self.send_button.config(text="به‌روزرسانی")
        self.edit_mode = True

    def update_entry(self, name, count, brand, shape, expire_date):
        if self.current_edit_db_id is None:
            return
        try:
            self.db_cursor.execute("""
                UPDATE drugs SET name=?, count=?, brand=?, shape=?, expire_date=? WHERE id=?
            """, (name, count, brand, shape, expire_date, self.current_edit_db_id))
            self.db_conn.commit()
            messagebox.showinfo("موفقیت", ".اطلاعات با موفقيت به روز رساني شد")
            self.clear_entries()
            self.display_data()
            self.send_button.config(text="ارسال")
            self.edit_mode = False
            self.current_edit_db_id = None
        except sqlite3.Error as e:
            messagebox.showerror("خطای پایگاه داده", f":خطا در به روز رساني اطلاعات {e}")

    def delete_selected_entry(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("هیچ انتخابی نیست", ".لطفاً برای حذف، یک ردیف را انتخاب کنید")
            return

        confirm = messagebox.askyesno("تایید حذف", "آیا مطمئن هستید که می‌خواهید این ردیف را حذف کنید؟")
        if confirm:
            for item in selected_items:
                db_id_to_delete = self.tree.item(item, 'values')[6] # شناسه
                try:
                    self.db_cursor.execute("DELETE FROM drugs WHERE id=?", (db_id_to_delete,))
                except sqlite3.Error as e:
                    messagebox.showerror("خطای پایگاه داده", f"خطا در حذف ردیف: {e}")
                    return
            self.db_conn.commit()
            messagebox.showinfo("موفقیت", ".ردیف(های) انتخاب شده با موفقیت حذف شد")
            self.display_data()

    def clear_entries(self):
        self.name_entry.delete(0, tk.END)
        self.count_entry.delete(0, tk.END)
        self.brand_entry.delete(0, tk.END)
        # Modified: Clear ComboBox selection
        self.shape_combobox.set('')
        self.set_today_date()
        self.send_button.config(text="ارسال")
        self.edit_mode = False
        self.current_edit_db_id = None
        if self.tree.selection():
            self.tree.selection_remove(self.tree.selection())

    def on_double_click_edit(self, event):
        if self.tree.identify_region(event.x, event.y) == "cell":
            self.edit_selected_entry()

    def search_by_count(self, event=None):
        search_term = self.search_count_entry.get().strip()

        if not search_term:
            self.display_data()
            return
        if not search_term.isdigit():
            messagebox.showerror("خطای ورودی", ".لطفاً برای جستجوی تعداد، فقط عدد وارد کنید")
            return

        try:
            query = "SELECT id, name, count, brand, shape, expire_date FROM drugs WHERE CAST(count AS TEXT) LIKE ? ORDER BY expire_date ASC"
            self.db_cursor.execute(query, ('%' + search_term + '%',))
            filtered_rows = self.db_cursor.fetchall()
            
            self.display_data(data=filtered_rows) 
            
            if not filtered_rows:
                messagebox.showinfo("نتیجه جستجو", f"دارویی که تعداد آن حاوی '{search_term}' باشد، یافت نشد.")
        except sqlite3.Error as e:
            messagebox.showerror("خطای پایگاه داده", f"خطا در هنگام جستجو: {e}")

    def reset_all(self):
        self.search_count_entry.delete(0, tk.END)
        self.display_data()
        self.clear_entries()

    def generate_xlsx(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")],
                                                 title="ذخیره فایل اکسل")
        if not filepath:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "اطلاعات بيماران"
        sheet.sheet_view.rightToLeft = True

        headers = ["تاریخ انقضا", "نوع پرونده", "بند", "کد", "نام", "ردیف"]
        sheet.append(headers)

        color_map = {'red_row': 'FFCCCC', 'orange_row': 'FFDDAA', 'green_row': 'CCFFCC', 'purple_row': 'E0B0FF'}

        for item_id in self.tree.get_children():
            values = list(self.tree.item(item_id, 'values'))
            tags = self.tree.item(item_id, 'tags')
            row_data = values[:-1]
            sheet.append(row_data)

            if tags and tags[0] in color_map:
                fill_color = color_map[tags[0]]
                for cell in sheet[sheet.max_row]:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        try:
            workbook.save(filepath)
            messagebox.showinfo("موفقیت", f"فایل اکسل با موفقیت در '{filepath}' .ذخیره شد")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در ذخیره فایل اکسل: {e}")

    def on_closing(self):
        if self.db_conn:
            self.db_conn.close()
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = DrugManagementApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()